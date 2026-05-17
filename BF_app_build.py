# 給与明細自動作成アプリ　テスト版

import os
import sys
import shutil
import tkinter as tk
import customtkinter as ctk
import gspread
from geopy.geocoders import Nominatim
import time # ネット検索に負荷をかけないための休憩用
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook
from datetime import datetime
import subprocess

# ==========================================
# 1. パス（住所）設定ゾーン
#    アプリが自分の「鍵」や「テンプレート」をどこから探すか決める場所
# ==========================================
def get_secure_path(filename):
    """
    【重要】Macのセキュリティ対策
    アプリ内部にあるファイルを、安全な『書類』フォルダにコピーして読み込む魔法
    """
    target_dir = os.path.expanduser("~/Documents/SalaryAppData")
    if not os.path.exists(target_dir):
        os.makedirs(target_dir)
    
    target_path = os.path.join(target_dir, filename)

    if hasattr(sys, 'frozen'):
        # アプリとして起動している時：中身を書類フォルダへコピー
        base_path = os.environ.get('RESOURCEPATH', os.path.dirname(sys.executable))
        source_path = os.path.join(base_path, filename)
        if os.path.exists(source_path):
            shutil.copy2(source_path, target_path)
    
    # 開発中（VS Code）なら、そのままカレントディレクトリのファイルを使う
    if not os.path.exists(target_path) and os.path.exists(filename):
        return os.path.abspath(filename)

    return target_path

# 実際に使うファイルの住所を確定
JSON_PATH     = get_secure_path('secret-key.json')
TEMPLATE_PATH = get_secure_path('template.xlsx')

# 出来上がったエクセルを保存するフォルダの名前を決めて、無ければ作る
SAVE_DIR = os.path.expanduser('~/Documents/給与明細保存先')
if not os.path.exists(SAVE_DIR):
    os.makedirs(SAVE_DIR)

# ==========================================
# 2. スプレッドシート・列の設定ゾーン
#    スプレッドシートの「どの列」に「何が」書いてあるかを定義する場所
# ==========================================
SPREADSHEET_ID = '1ab6ck_N4Kue2z3QJazx6L86etmzcoPJNPA7WcwlNQ2Y'
COL_NAME  = '名前 例:天江拓也(スペース無し)'
COL_START = '稼働開始日'
COL_END   = '稼働終了日'
COL_SHOP  = '稼働店舗 例:コジマ盛岡'
COL_QTY   = '数量'
COL_EXPENSE_DETAIL  = '立替金額(合計)'
COL_TYPE  = '案件タイプ'
COL_REPORT_TYPE = '報告種別(ディレクターのみ回答)' 

# --- Googleへのログイン（認証） ---
try:
    scopes = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
    credentials = Credentials.from_service_account_file(JSON_PATH, scopes=scopes)
    gc = gspread.authorize(credentials)
except Exception as e:
    print(f"認証エラー: {e}")

# ==========================================
# 3. アプリの見た目（画面）ゾーン
#    ボタンや入力欄を作っている場所
# ==========================================
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("給与精算システム")
        self.geometry("400x450")
        
        font_main = ("Hiragino Sans", 14)
        font_title = ("Hiragino Sans", 20, "bold")

        self.label = ctk.CTkLabel(self, text="給与明細 自動作成", font=font_title)
        self.label.pack(pady=20)
        
        self.name_entry = ctk.CTkEntry(self, placeholder_text="名前を入力", width=250, font=font_main)
        self.name_entry.pack(pady=10)
        
        self.month_entry = ctk.CTkEntry(self, placeholder_text="月を入力", width=250, font=font_main)
        self.month_entry.pack(pady=10)
        
        # ボタンが押されたら「button_callback」を実行せよという命令
        self.button = ctk.CTkButton(self, text="明細書を作成する", command=self.button_callback, font=font_main)
        self.button.pack(pady=20)

    # ==========================================
    # 4. メインロジック（中身の計算）ゾーン
    # ==========================================
    def button_callback(self):
        target_name = self.name_entry.get().replace(" ", "").replace("　", "").strip()
        month_input = self.month_entry.get().lstrip('0').strip()
        
        if not target_name or not month_input:
            self.label.configure(text="名前と月を入力してください", text_color="red")
            return

        m_num = month_input
        m_with_zero = m_num.zfill(2)

        self.label.configure(text="データ抽出中...", text_color="orange")
        self.update()

        try:
            # --- 4-A. スプレッドシートからデータを全部持ってくる ---
            sh = gc.open_by_key(SPREADSHEET_ID)

            def safe_get_records(worksheet_obj):
                all_rows = worksheet_obj.get_all_values()
                if not all_rows:
                    return []
                raw_headers = all_rows[0]
                valid_headers = [(i, h) for i, h in enumerate(raw_headers) if h.strip() != ""]
                
                result_data = []
                for row_cells in all_rows[1:]:
                    if len(row_cells) < len(raw_headers):
                        row_cells += [""] * (len(raw_headers) - len(row_cells))
                    row_dict = {}
                    for idx, header_name in valid_headers:
                        row_dict[header_name] = row_cells[idx]
                    result_data.append(row_dict)
                return result_data

            ans_data     = safe_get_records(sh.get_worksheet(0))         
            staff_list   = safe_get_records(sh.worksheet("スタッフ名簿")) 
            price_master = safe_get_records(sh.worksheet("単価マスタ"))   

            # --- 4-B. その人の基本「ランク」と「単価」を特定する ---
            staff_rank = next((row['ランク'] for row in staff_list if str(row['氏名']).replace(" ", "").replace("　", "") == target_name), None)
            
            raw_price = next((row['単価'] for row in price_master if str(row['ランク']) == staff_rank), 0)
            unit_price = int(str(raw_price).replace(",", "") or 0)

            telecom_jobs = []   
            other_jobs = []     
            total_expenses_sum = 0     
            expense_details_log = []    

            # --- 4-C. 全データの中から「今回の人・月」だけを抜き出すループ ---
            for row in ans_data:
                name_in_row = str(row.get(COL_NAME, "")).replace(" ", "").replace("　", "")
                if name_in_row != target_name:
                    continue

                date_str = str(row.get(COL_START, ""))
                if not (f"/{m_num}/" in date_str or f"/{m_with_zero}/" in date_str):
                    continue

                try:
                    clean_date = date_str.split()[0]
                    parts = clean_date.split('/')
                    date_label = f"{parts[1].lstrip('0')}/{parts[2].lstrip('0')}"
                    
                    end_val = str(row.get(COL_END, ""))
                    if end_val and end_val != date_str:
                        e_parts = end_val.split()[0].split('/')
                        if len(e_parts) >= 3:
                            date_label += f"-{e_parts[2].lstrip('0')}"
                    
                    dt_obj = datetime.strptime(clean_date, '%Y/%m/%d')
                except:
                    date_label = date_str
                    dt_obj = datetime.now()

                shop_name_str = str(row.get(COL_SHOP, "")).strip()
                display_text = f"{date_label} {shop_name_str}"
                
                # 立替金（実費請求）の計算
                try:
                    r_val = str(row.get(COL_EXPENSE_DETAIL, "0")).replace(",", "")
                    reimb_amt = float(r_val or 0)
                except: 
                    reimb_amt = 0
                
                if reimb_amt > 0:
                    amt_int = int(reimb_amt)
                    total_expenses_sum += amt_int 
                    log_line = f"・{date_label} {shop_name_str}: 実費 ({amt_int:,}円)"
                    expense_details_log.append(log_line)

                # 手当判定処理
                job_type = str(row.get(COL_TYPE, "")) 
                shop_name = shop_name_str 
                report_val = str(row.get(COL_REPORT_TYPE, "")).strip() 

                item = {
                    'date': dt_obj,
                    'text': display_text,
                    'quantity': float(str(row.get(COL_QTY, "1")).replace(",", "") or 1),
                    'price': unit_price, 
                    'type': job_type
                }

                if "通信" not in job_type:
                    item['price'] = None  
                    other_jobs.append(item) 
                    continue

                if staff_rank in ["コアA", "コアS"]:
                    telecom_jobs.append(item)  
                    continue
                
                calculated_price = unit_price

                # ① 県外判定
                if shop_name:
                    try:
                        geolocator = Nominatim(user_agent="SalaryApp_Takuya")
                        location = geolocator.geocode(f"{shop_name}, Japan", timeout=5)
                        time.sleep(0.5) 
                        
                        is_outside = False
                        if location and location.address:
                            if "宮城" not in location.address:
                                is_outside = True
                        else:
                            if not any(k in shop_name for k in ["宮城", "仙台", "名取"]):
                                is_outside = True
                                
                        if is_outside:
                            calculated_price += 2000 
                            
                    except Exception as geo_err:
                        print(f"住所検索エラー（スキップ）: {geo_err}")

                # ② ディレクター手当の判定
                is_director = report_val and report_val not in ["両方なし", "nan", ""]
                if is_director:
                    calculated_price += 2000 

                item['price'] = calculated_price
                telecom_jobs.append(item)

                # ③ 報告手当
                report_fee = 0
                if "報告書のみ" in report_val:
                    report_fee = 3000
                elif "報告書+報告会" in report_val or "両方" in report_val:
                    report_fee = 5000

                if report_fee > 0:
                    other_jobs.append({
                        'date': dt_obj,
                        'text': f"{date_label} ディレクター報告手当",
                        'quantity': 1,
                        'price': report_fee, 
                        'type': "手当"
                    })

            # 日付順に並び替える
            telecom_jobs.sort(key=lambda x: x['date'])
            other_jobs.sort(key=lambda x: x['date'])

            # グループ化して集計
            aggregated_telecom = {}
            for job in telecom_jobs:
                shop_pure_name = str(job['text'].split(maxsplit=1)[1]) if " " in job['text'] else job['text']
                key = (shop_pure_name, job['price'])
                if key not in aggregated_telecom:
                    aggregated_telecom[key] = {
                        'shop_name': shop_pure_name,
                        'price': job['price'],
                        'raw_jobs': [] 
                    }
                aggregated_telecom[key]['raw_jobs'].append(job)

            if not telecom_jobs and not other_jobs:
                self.label.configure(text=f"データが見つかりませんでした", text_color="red")
                return

            # --------------------------------------------------
            # --- 4-D. エクセルへの書き出しゾーン（完全修正版） ---
            # --------------------------------------------------
            wb = load_workbook(TEMPLATE_PATH)
            ws = wb.active
            ws['A1'].value = f"外注明細書:{target_name}({m_num}月)"

            # 結合セルの罠を回避して安全に書き込む関数
            def safe_write(worksheet, coord, val):
                from openpyxl.cell import MergedCell
                target_cell = worksheet[coord]
                if isinstance(target_cell, MergedCell):
                    for m_range in worksheet.merged_cells.ranges:
                        if target_cell.coordinate in m_range:
                            worksheet.cell(row=m_range.min_row, column=m_range.min_col).value = val
                            return
                else:
                    target_cell.value = val

            # 通信案件を10行目から順に書く
            row_idx = 10
            for (shop_pure_name, price), info in aggregated_telecom.items():
                
                first_date_text = info['raw_jobs'][0]['text'].split()[0]
                last_date_text = info['raw_jobs'][-1]['text'].split()[0]
                
                total_days = 0
                for job in info['raw_jobs']:
                    date_str = job['text'].split()[0]
                    if "-" in date_str:
                        try:
                            parts = date_str.split('/')[-1].split('-')
                            total_days += (int(parts[1]) - int(parts[0]) + 1)
                        except: total_days += 1
                    else:
                        total_days += 1
                
                if first_date_text == last_date_text:
                    date_label_final = first_date_text
                else:
                    date_label_final = f"{first_date_text.split('-')[0]}-{last_date_text.split('-')[-1]}"
                
                clean_rank = str(staff_rank).replace("+", "").strip()
                
                # ★修正1: セルごと改行（行を分離）
                # 1行目: 日付と店舗名
                shop_line_text = f"{date_label_final} {info['shop_name']}"
                safe_write(ws, f'B{row_idx}', shop_line_text)
                
                # 2行目: 現場単価と金額データ（元の手書きフォーマットと完全一致！）
                rank_line_text = f"現場単価：{clean_rank}"
                safe_write(ws, f'B{row_idx+1}', rank_line_text)
                safe_write(ws, f'H{row_idx+1}', info['price'])     # H列：単価
                safe_write(ws, f'I{row_idx+1}', total_days)        # I列：数量（日数）
                safe_write(ws, f'J{row_idx+1}', "日")              # J列：単位
                
                # 3行目を完全な空行にしてメリハリをつけるため、次の開始行を「3」進める
                row_idx += 3

            # その他案件（研修・報告手当など）を少し空けて書く
            row_idx += 1
            for j in other_jobs:
                safe_write(ws, f'B{row_idx}', j['text'])
                safe_write(ws, f'H{row_idx}', j.get('price'))    
                safe_write(ws, f'I{row_idx}', j['quantity'])     
                safe_write(ws, f'J{row_idx}', "")                
                row_idx += 1

            # ★修正2: 40行目の文字を「立替金合計」のみにスッキリ修正
            if total_expenses_sum > 0:
                safe_write(ws, 'B40', "立替金合計") # B40：余計な文字を排除
                safe_write(ws, 'H40', total_expenses_sum)               # H40：金額
                safe_write(ws, 'I40', 1)                                # I40：数量   
                safe_write(ws, 'J40', "回")                             # J40：単位

            # --- 4-E. 保存ゾーン ---
            save_filename = f"{datetime.now().year}年{m_num}月給与明細_{target_name}.xlsx"
            save_full_path = os.path.join(SAVE_DIR, save_filename)
            wb.save(save_full_path)

            # ★【ここを追加】保存完了後、Macで自動的にFinderを開く
            subprocess.run(["open", SAVE_DIR])
            
            if expense_details_log:
                log_text = f"【{target_name}様の立替金内訳】\n" + "\n".join(expense_details_log)
            else:
                log_text = f"【{target_name}様の立替金内訳】\n・今月の立替金請求はありません。"

            # アプリ画面に内訳表示用のテキストボックスを生成
            import tkinter as tk
            from tkinter import scrolledtext
            
            if hasattr(self, 'expense_box'):
                self.expense_box.destroy()
            
            self.expense_box = scrolledtext.ScrolledText(self, width=50, height=6, bg="#2b2b2b", fg="#ffffff", font=("Arial", 11))
            self.expense_box.pack(pady=10)
            self.expense_box.insert(tk.END, log_text)
            self.expense_box.configure(state='disabled')

            self.label.configure(text=f"作成完了！\n書類フォルダを確認してください", text_color="green")

        except Exception as e:
            self.label.configure(text="エラーが発生しました", text_color="red")
            print(f"Error details: {e}")

# ==========================================
# 5. 起動スイッチ
# ==========================================
if __name__ == "__main__":
    app = App()
    app.mainloop()